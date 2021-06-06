import pathlib
import os
import logging
import openpyxl

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

logging.info("change path for current file dir ")
os.chdir(pathlib.Path(
    __file__).parent.absolute())

excel_my: str = str(pathlib.Path(
    __file__).parent.absolute()) + "/MOCK_DATA.xlsx"


def read_fro_excel():
    logging.info("read from excel start")

    workbook = openpyxl.load_workbook(excel_my)
    # print(workbook.sheetnames)
    sheet_ = workbook['data']

    print(sheet_["c"][:5])

    for i in sheet_["c"][:5]:
        print(i.value, end=" , ")
    print()

    for i in range(1, 5):
        print(sheet_.cell(row=i, column=2).value)


def edit_excel_file():
    logging.info("create new wb")
    wb = openpyxl.Workbook()

    print(wb.sheetnames)
    print(wb['Sheet'])
    sheet_ = wb['Sheet']
    print(sheet_.cell(1, 1).value)  # none 

    wb.save('my_book.xlsx')


if __name__ == "__main__":
    # read_fro_excel()
    edit_excel_file()
